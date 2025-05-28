import streamlit as st
import zipfile
import time
import warnings
from pathlib import Path
from io import BytesIO
from openpyxl import load_workbook
import pandas as pd

warnings.filterwarnings(
    "ignore",
    message="File contains an invalid specification for 0. This will be removed"
)

st.set_page_config(
    page_title="SPDO Resumo de Fichas Técnicas",
    page_icon="logo_fgv.png",
    layout="wide"
)
st.logo('logo_ibre.png')
st.title("Resumo de Fichas Técnicas")
st.markdown("Faça upload da pasta **Fichas Tecnicas** compactada em .zip para iniciar o processamento automaticamente.")
uploaded_zip = st.file_uploader("Upload da pasta Fichas Tecnicas (zip)", type=["zip"])
if uploaded_zip:
    if 'df_records' not in st.session_state:
        try:
            with zipfile.ZipFile(uploaded_zip) as zf:
                members = [n for n in zf.namelist() if n.lower().endswith(('.xlsx','.xls'))]
                total = len(members)
                if total == 0:
                    st.error("Nenhum arquivo .xlsx encontrado dentro do ZIP.")
                    st.stop()
                pc, tx = st.columns([1,4])
                pb = pc.progress(0)
                stx = tx.empty()
                recs = []
                start = time.perf_counter()
                for i, m in enumerate(members,1):
                    pb.progress(i/total)
                    stx.text(f"Processando {i}/{total}: {m}")
                    parts = Path(m).parts
                    if len(parts)>=3:
                        f,s=parts[0],parts[1]
                        fn=parts[-1]
                    elif len(parts)==2:
                        f,fn=parts
                        s=""
                    else:
                        f,s,fn="","",parts[-1]
                    data=zf.read(m)
                    wb=load_workbook(filename=BytesIO(data),data_only=True)
                    ws=wb.active
                    stem=Path(fn).stem
                    if s=="CAGECE":
                        c,a,e="D5","H5","K9"
                    elif s=="DER-MG":
                        if stem=="MATRO-1794":
                            c,a,e="D3","I3","N9"
                        elif stem in ["EQRO-1508","EQRO-5651","EQRO-5652"]:
                            c,a,e="D4","H4","N9"
                        else:
                            c,a,e="D5","H5","K9"
                    elif s=="ECON-DNIT":
                        if stem in ["E8888", "E8351", "E8306"]:
                            c,a,e ="D4", "H4", "N9"
                        elif stem in ["M7062", "M7099", "M7228", "M7618", "M7642"]:
                            c,a,e ="D3", "I3", "N9"
                        else:
                            c,a,e = "D5", "H5", "K7"
                    elif s=="SANEAGO":
                        c,a,e ="D3", "I3", "N9"
                    elif s=="SICRO":
                        if stem in [
                            "M0291", "M0292", "M0293", "M0294", "M0295", "M0296", "M0297",
                            "M0375", "M0713", "M0714", "M0715", "M0716", "M1728", "M1729",
                            "M1730", "M1894", "M1895", "M1899", "M1920", "M1923", "M2017",
                            "M2019", "M2020", "M2029", "M2030", "M2031", "M2032", "M2033",
                            "M2035", "M2089", "M2090", "M2091", "M2094", "M2095", "M2096",
                            "M2101", "M2325", "M2326", "M2327", "M2328", "M2329", "M2330",
                            "M2331", "M2332", "M2333", "M2585", "M2586", "M2587", "M2588",
                            "M2589", "M2590", "M2591", "M3091", "M3094", "M3523", "M3825",
                            "M3829", "M3843", "M3853"
                        ]:
                            c,a,e = "D3", "I3", "N9"
                        elif stem in [
                            "A9316", "A9344", "A9345", "E9066", "E9145", "E9259", "E9260",
                            "E9261", "E9262", "E9263", "E9264", "E9265", "E9267", "E9543",
                            "E9575", "E9579", "E9667"
                        ]:
                            c,a,e = "D4", "H4", "N9"
                        else:
                            c,a,e = "D5", "H5", "K9"
                    else:
                        c,a,e="D4","H4","K9"
                    def sc(ws,cr,d="-"):
                        try:
                            v=ws[cr].value
                            return d if v is None else v
                        except:
                            return d
                    r={"Pasta":f,"Subpasta":s,"Arquivo":fn}
                    r[c]=sc(ws,c)
                    r[a]=sc(ws,a)
                    r[e]=sc(ws,e)
                    recs.append(r)
            elapsed=time.perf_counter()-start
            st.success(f"Processamento concluído em {elapsed:.2f} segundos!")
            zf.close()
            st.session_state['df_records'] = pd.DataFrame(recs)
            if recs:
                df=pd.DataFrame(recs)
                df = df.drop(columns=['Pasta'])
                df = df.rename(columns={
                    'Subpasta': 'Contrato',
                    'Arquivo': 'Nome do Arquivo',
                    'D5': 'Data Criação1',
                    'H5': 'Data Atualização1',
                    'K9': 'Codigo1',
                    'D3': 'Data Criação2',
                    'I3': 'Data Atualização2',
                    'N9': 'Codigo2',
                    'D4': 'Data Criação3',
                    'H4': 'Data Atualização3',
                    'K7': 'Codigo3',
                })

                date_cols = [
                    'Data Criação1','Data Criação2','Data Criação3',
                    'Data Atualização1','Data Atualização2','Data Atualização3'
                ]
                for c in date_cols:
                    # converte para datetime, depois formata. valores não convertidos viram NaT → string "-"
                    df[c] = (
                        pd.to_datetime(df[c], errors='coerce')
                        .dt.strftime('%d/%m/%Y')
                        .fillna('-')
                    )
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Dados')
                output.seek(0)


                df = st.session_state['df_records']
                st.subheader("Dados Extraídos")
                st.dataframe(df)
                
                st.download_button(
                    "📥 Baixar Excel",
                    data=output,
                    file_name="dados_extraidos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.info("Nenhum registro extraído dos arquivos.")
        except zipfile.BadZipFile:
            st.error("O arquivo enviado não é um ZIP válido.")
